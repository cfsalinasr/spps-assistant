"""spps-assistant template — generate blank materials input templates."""

from pathlib import Path

import click
from rich.console import Console
from rich.panel import Panel

console = Console()


@click.command('template')
@click.option('--output-dir', '-o', default='.', type=click.Path(),
              help='Directory to write template files (default: current directory).')
def template(output_dir: str) -> None:
    """Generate blank materials input template files.

    Creates:
        spps_materials_template.csv   — CSV template for reactant information
        spps_materials_template.xlsx  — XLSX version for spreadsheet editing
        spps_sequences_template.fasta — FASTA template for peptide sequences

    Fill these in and pass to --materials / --input when running 'generate' or 'materials'.
    """
    import csv
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    headers = [
        'ResidueCode', 'ProtectionGroup', 'FmocMW_g_mol',
        'FreeAA_MW_g_mol', 'Density_g_mL', 'Equivalents', 'Notes'
    ]

    example_rows = [
        # Amino acids (solid — leave Density_g_mL blank; Equivalents = 1 × reactant excess)
        ['A',     '',    '311.3', '71.08',  '',      '1',  'Fmoc-Ala-OH'],
        ['G',     '',    '297.3', '57.05',  '',      '1',  'Fmoc-Gly-OH'],
        ['C',     'Trt', '585.7', '103.14', '',      '1',  'Fmoc-Cys(Trt)-OH'],
        ['C',     'Acm', '446.5', '103.14', '',      '1',  'Fmoc-Cys(Acm)-OH — orthogonal, NOT removed by TFA'],
        ['K',     'Boc', '468.6', '128.17', '',      '1',  'Fmoc-Lys(Boc)-OH'],
        ['R',     'Pbf', '648.8', '156.19', '',      '1',  'Fmoc-Arg(Pbf)-OH'],
        # Activators — solid (Equivalents = 1 × reactant excess)
        ['HBTU',  '',    '379.3', '379.3',  '',      '1',  'Coupling activator — solid'],
        ['OXYMA', '',    '142.1', '142.1',  '',      '1',  'Coupling additive — solid'],
        # Liquid reagents — Equivalents multiplier relative to reactant excess
        # DIEA: 2× the reactant excess (e.g. excess=10 → 20 eq total)
        ['DIEA',  '',    '129.24','129.24', '0.742', '2',  'Base — liquid; 2× reactant excess'],
        # DIC: same as reactant excess
        ['DIC',   '',    '126.2', '126.2',  '0.806', '1',  'Coupling activator — liquid'],
        # Pyridine: 20× the reactant excess (e.g. excess=10 → 200 eq total)
        ['Pyridine', '', '79.1', '79.1',   '0.978', '20', 'Catalyst — liquid; 20× reactant excess'],
    ]

    # CSV template
    csv_path = out_path / 'spps_materials_template.csv'
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(example_rows)
        for _ in range(20):
            writer.writerow([''] * len(headers))

    console.print(f"  [green]Reactants CSV template:[/green]  {csv_path}")

    # XLSX template
    xlsx_path = out_path / 'spps_materials_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materials'

    thin_side = Side(style='thin', color='AAAAAA')
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    col_widths = [14, 18, 18, 18, 14, 14, 45]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    for row_idx, row_data in enumerate(example_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=12)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='EBF5FB')

    for row_idx in range(len(example_rows) + 2, len(example_rows) + 22):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'
    wb.save(str(xlsx_path))

    console.print(f"  [green]Reactants XLSX template:[/green] {xlsx_path}")

    # Sequence templates
    fasta_path = out_path / 'spps_sequences_template.fasta'
    fasta_content = (
        "; SPPS Sequence Input Template — FASTA format\n"
        "; Lines starting with ';' are comments and are ignored.\n"
        "; Each peptide starts with a '>' header line followed by the sequence.\n"
        "; Use bracket notation for protected residues, e.g. C(Trt), K(Boc), C(Acm).\n"
        "; NOTE: C(Acm) carries an orthogonal protecting group not removed by TFA —\n"
        ";       a separate post-synthesis deprotection step is required.\n"
        ";\n"
        ">Peptide_1\n"
        "AC(Trt)DEFGHIK(Boc)\n"
        ">Peptide_2\n"
        "GLC(Acm)MVWS(tBu)\n"
        ">Peptide_3\n"
        "ALYQK(Boc)VFANIK(Boc)\n"
    )
    fasta_path.write_text(fasta_content, encoding='utf-8')
    console.print(f"  [green]FASTA sequence template:[/green] {fasta_path}")

    console.print(Panel(
        "[bold]How to use the templates:[/bold]\n\n"
        "[bold]1.[/bold] Fill in [bold]spps_materials_template.xlsx[/bold] with information\n"
        "   about your reactants (amino acids, activators, bases).\n"
        "   • [bold]Equivalents[/bold]: multiplier relative to your reactant excess.\n"
        "     Regular AAs = 1, DIEA = 2, Pyridine = 20.\n"
        "   • Leave [bold]Density_g_mL[/bold] blank for solid reagents.\n"
        "   • Fill in [bold]Density_g_mL[/bold] for liquid reagents (e.g. DIEA, DIC).\n"
        "     Liquid quantities are reported in µL (no weighing needed).\n\n"
        "[bold]2.[/bold] Fill in [bold]spps_sequences_template.fasta[/bold] with your peptide\n"
        "   sequences. Use bracket notation for protected residues:\n"
        "   C(Trt), K(Boc), C(Acm), S(tBu) ...\n\n"
        "[bold]3.[/bold] In the launcher, choose:\n"
        "   • [bold]Option 1[/bold] — Generate synthesis guide (PDF + DOCX)\n"
        "   • [bold]Option 2[/bold] — Generate materials report (XLSX)\n"
        "   A Finder window will open so you can select your files.\n\n"
        "Accepted formats: reactants (.xlsx or .csv) · sequences (.fasta)",
        title="Input Templates",
        border_style="cyan",
    ))
