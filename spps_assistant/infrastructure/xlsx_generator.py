"""openpyxl XLSX generator for SPPS materials lists."""

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from spps_assistant.domain.models import MaterialsRow

# Color palette
COLOR_HEADER_BG = '2C3E50'
COLOR_HEADER_FG = 'FFFFFF'
COLOR_ALT_ROW    = 'EBF5FB'
COLOR_WHITE      = 'FFFFFF'
COLOR_ACCENT     = '1A5276'


def _thin_border() -> Border:
    """Return a thin grey border for all four sides of a cell."""
    side = Side(style='thin', color='AAAAAA')
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill() -> PatternFill:
    """Return the dark-blue solid fill used for header rows."""
    return PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)


def _alt_fill() -> PatternFill:
    """Return the light-blue solid fill used for alternating data rows."""
    return PatternFill(fill_type='solid', fgColor=COLOR_ALT_ROW)


def _apply_cell_number_format(cell, col_idx: int, is_liquid: bool) -> None:
    """Apply number format for numeric columns in the materials XLSX."""
    if col_idx == 3:
        cell.number_format = '0.0'
    elif col_idx == 4:
        cell.number_format = '0.0000'
    elif col_idx == 5 and not is_liquid:
        cell.number_format = '0.00'
    elif col_idx == 6:
        cell.number_format = '0.00'
    elif col_idx == 7:
        cell.number_format = '0.0 "µL"' if is_liquid else '0.000'


def _write_mat_row(ws, row_offset: int, mat: MaterialsRow) -> None:
    """Write a single materials row to the worksheet."""
    row_num = row_offset + 3
    is_alt = row_offset % 2 == 1
    is_liquid = mat.volume_ul is not None

    if is_liquid:
        mass_cell_value = None
        volume_cell_value = mat.volume_ul
    else:
        mass_cell_value = mat.mass_mg
        volume_cell_value = mat.volume_ml

    row_data = [
        mat.token, mat.protection, mat.fmoc_mw, mat.mmol_needed,
        mass_cell_value, mat.stock_conc, volume_cell_value,
        mat.formula, mat.notes,
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name='Calibri', size=12)
        cell.border = _thin_border()
        cell.alignment = Alignment(
            horizontal='center' if col_idx <= 7 else 'left',
            vertical='center',
            wrap_text=(col_idx >= 8),
        )
        if is_liquid:
            cell.fill = PatternFill(fill_type='solid', fgColor='FEF9E7')
        elif is_alt:
            cell.fill = _alt_fill()
        _apply_cell_number_format(cell, col_idx, is_liquid)

    ws.row_dimensions[row_num].height = 22


def generate_materials_xlsx(
    path: Path,
    synthesis_name: str,
    materials_rows: List[MaterialsRow],
) -> None:
    """Generate a bench-optimized materials list XLSX.

    Layout:
        - Large font (14pt), bold headers
        - Color-coded alternating rows
        - Columns: Residue | Protection | Fmoc-MW (g/mol) | mmol needed |
                   Mass to weigh (mg) | Stock Conc (M) | Volume (mL) |
                   Formula | Notes

    Args:
        path: Output XLSX file path
        synthesis_name: Synthesis run name (used in sheet title)
        materials_rows: List of MaterialsRow objects
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = synthesis_name[:31]  # Sheet name max 31 chars

    # ------------------------------------------------------------------ #
    # Title row                                                            #
    # ------------------------------------------------------------------ #
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"Materials List — {synthesis_name}"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color=COLOR_HEADER_FG)
    title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_ACCENT)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ------------------------------------------------------------------ #
    # Header row (row 2)                                                   #
    # ------------------------------------------------------------------ #
    headers = [
        'Residue', 'Protection', 'Fmoc-MW (g/mol)', 'mmol needed',
        'Mass (mg) / — liquid', 'Stock Conc (M)', 'Volume (mL / µL)',
        'Formula', 'Notes',
    ]
    col_widths = [12, 14, 18, 14, 20, 16, 16, 45, 30]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER_FG)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # ------------------------------------------------------------------ #
    # Data rows                                                            #
    # ------------------------------------------------------------------ #
    for row_offset, mat in enumerate(materials_rows):
        _write_mat_row(ws, row_offset, mat)

    # ------------------------------------------------------------------ #
    # Freeze panes below header                                            #
    # ------------------------------------------------------------------ #
    ws.freeze_panes = 'A3'

    # ------------------------------------------------------------------ #
    # Auto-filter on headers                                               #
    # ------------------------------------------------------------------ #
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(materials_rows) + 2}"

    wb.save(str(path))
