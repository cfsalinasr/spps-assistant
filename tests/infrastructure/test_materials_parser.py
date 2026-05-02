"""Tests for infrastructure/materials_parser.py."""

import csv
from pathlib import Path

import openpyxl
import pytest

from spps_assistant.infrastructure.materials_parser import (
    _parse_float,
    load_materials_file,
    parse_materials_csv,
    parse_materials_xlsx,
)


# ── _parse_float ──────────────────────────────────────────────────────────────

class TestParseFloat:
    def test_dot_decimal(self):
        """Dot decimal separator is parsed correctly."""
        assert _parse_float('311.3', 0.0) == pytest.approx(311.3)

    def test_comma_decimal(self):
        """Comma decimal separator is normalised to dot."""
        assert _parse_float('311,3', 0.0) == pytest.approx(311.3)

    def test_empty_string_returns_default(self):
        """Empty string returns the default value."""
        assert _parse_float('', 99.0) == pytest.approx(99.0)

    def test_none_returns_default(self):
        """None input returns the default value."""
        assert _parse_float(None, 5.0) == pytest.approx(5.0)

    def test_invalid_string_returns_default(self):
        """Non-numeric string returns the default value."""
        assert _parse_float('abc', 1.0) == pytest.approx(1.0)

    def test_integer_string(self):
        """Integer string is parsed as a float."""
        assert _parse_float('42', 0.0) == pytest.approx(42.0)

    def test_zero(self):
        """String '0' parses to 0.0, not the default."""
        assert _parse_float('0', 99.0) == pytest.approx(0.0)

    def test_native_float_passthrough(self):
        """Native float is returned unchanged."""
        assert _parse_float(379.3, 0.0) == pytest.approx(379.3)


# ── parse_materials_csv ───────────────────────────────────────────────────────

class TestParseMaterialsCSV:
    def test_basic_parsing(self, tmp_path):
        """Standard CSV is parsed into the expected token dict."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311.3,71.08,,Fmoc-Ala-OH\n'
            'C,Trt,585.7,103.14,,Fmoc-Cys(Trt)-OH\n'
        )
        rows = parse_materials_csv(f)
        assert len(rows) == 2
        assert rows[0]['token'] == 'A'
        assert rows[0]['base_code'] == 'A'
        assert rows[0]['protection'] == ''
        assert rows[0]['fmoc_mw'] == pytest.approx(311.3)
        assert rows[1]['token'] == 'C(Trt)'
        assert rows[1]['protection'] == 'Trt'

    def test_comma_decimal_separator(self, tmp_path):
        """Comma decimal in a field is handled by _parse_float."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311,3,71,08,,\n'
        )
        # Comma-delimited CSV where decimals use commas — handled by _parse_float
        f2 = tmp_path / 'mat2.csv'
        f2.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'DIEA,,129,24,129,24,0,742,Base\n'
        )
        # The decimal comma within a field (when CSV uses , as field separator is ambiguous)
        # Test the _parse_float behaviour directly on the value
        assert _parse_float('129,24', 0.0) == pytest.approx(129.24)

    def test_density_parsed_for_liquid(self, tmp_path):
        """Density_g_mL column is parsed for liquid reagents."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'DIEA,,129.24,129.24,0.742,Base liquid\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['density_g_ml'] == pytest.approx(0.742)

    def test_density_none_for_solid(self, tmp_path):
        """Empty density cell yields None for solid reagents."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'HBTU,,379.3,379.3,,Solid activator\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['density_g_ml'] is None

    def test_empty_rows_skipped(self, tmp_path):
        """Rows with no ResidueCode are silently skipped."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311.3,71.08,,\n'
            ',,,,, \n'
            ',,,,,\n'
            'G,,297.3,57.05,,\n'
        )
        rows = parse_materials_csv(f)
        assert len(rows) == 2

    def test_backward_compat_stock_conc_column(self, tmp_path):
        """Legacy StockConc_M column is still accepted."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,StockConc_M,Notes\n'
            'A,,311.3,71.08,0.5,\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['stock_conc'] == pytest.approx(0.5)

    def test_file_not_found(self):
        """Non-existent CSV path raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            parse_materials_csv(Path('/nonexistent/file.csv'))

    def test_non_standard_activator_token(self, tmp_path):
        """Non-amino-acid tokens such as HBTU are preserved."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'HBTU,,379.3,379.3,,Coupling activator\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['token'] == 'HBTU'
        assert rows[0]['base_code'] == 'HBTU'

    def test_missing_fmoc_mw_defaults_to_zero(self, tmp_path):
        """Missing Fmoc MW defaults to 0.0."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,,,,\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['fmoc_mw'] == pytest.approx(0.0)

    def test_equivalents_multiplier_parsed(self, tmp_path):
        """Equivalents column is parsed into equivalents_multiplier."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Equivalents,Notes\n'
            'A,,311.3,71.08,,1,Fmoc-Ala-OH\n'
            'DIEA,,129.24,129.24,0.742,2,Base liquid\n'
            'Pyridine,,79.1,79.1,0.978,20,Catalyst\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['equivalents_multiplier'] == pytest.approx(1.0)
        assert rows[1]['equivalents_multiplier'] == pytest.approx(2.0)
        assert rows[2]['equivalents_multiplier'] == pytest.approx(20.0)

    def test_equivalents_multiplier_defaults_to_one_when_missing(self, tmp_path):
        """Missing Equivalents column defaults multiplier to 1.0."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311.3,71.08,,Fmoc-Ala-OH\n'
        )
        rows = parse_materials_csv(f)
        assert rows[0]['equivalents_multiplier'] == pytest.approx(1.0)


# ── parse_materials_xlsx ──────────────────────────────────────────────────────

class TestParseMaterialsXLSX:
    def _make_xlsx(self, tmp_path, rows):
        """Write a minimal XLSX materials file and return its path."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['ResidueCode', 'ProtectionGroup', 'FmocMW_g_mol',
                   'FreeAA_MW_g_mol', 'Density_g_mL', 'Notes'])
        for row in rows:
            ws.append(row)
        p = tmp_path / 'mat.xlsx'
        wb.save(str(p))
        return p

    def test_basic_xlsx_parsing(self, tmp_path):
        """Standard XLSX is parsed into the expected token dicts."""
        p = self._make_xlsx(tmp_path, [
            ['A', '', 311.3, 71.08, None, 'Fmoc-Ala-OH'],
            ['C', 'Trt', 585.7, 103.14, None, 'Fmoc-Cys(Trt)-OH'],
        ])
        rows = parse_materials_xlsx(p)
        assert len(rows) == 2
        assert rows[0]['token'] == 'A'
        assert rows[1]['token'] == 'C(Trt)'

    def test_density_from_xlsx(self, tmp_path):
        """Density cell in XLSX is read correctly."""
        p = self._make_xlsx(tmp_path, [
            ['DIEA', '', 129.24, 129.24, 0.742, 'Liquid base'],
        ])
        rows = parse_materials_xlsx(p)
        assert rows[0]['density_g_ml'] == pytest.approx(0.742)

    def test_empty_xlsx_returns_empty_list(self, tmp_path):
        """XLSX with no data rows returns an empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        p = tmp_path / 'empty.xlsx'
        wb.save(str(p))
        rows = parse_materials_xlsx(p)
        assert rows == []

    def test_xlsx_skips_empty_rows(self, tmp_path):
        """Blank rows in XLSX are skipped."""
        p = self._make_xlsx(tmp_path, [
            ['A', '', 311.3, 71.08, None, ''],
            [None, None, None, None, None, None],
            ['G', '', 297.3, 57.05, None, ''],
        ])
        rows = parse_materials_xlsx(p)
        assert len(rows) == 2

    def test_file_not_found(self):
        """Non-existent XLSX path raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            parse_materials_xlsx(Path('/nonexistent/file.xlsx'))


# ── load_materials_file ───────────────────────────────────────────────────────

class TestLoadMaterialsFile:
    def test_routes_csv(self, tmp_path):
        """CSV extension routes to CSV parser."""
        f = tmp_path / 'mat.csv'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311.3,71.08,,\n'
        )
        rows = load_materials_file(f)
        assert len(rows) == 1

    def test_routes_xlsx(self, tmp_path):
        """XLSX extension routes to XLSX parser."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['ResidueCode', 'ProtectionGroup', 'FmocMW_g_mol',
                   'FreeAA_MW_g_mol', 'Density_g_mL', 'Notes'])
        ws.append(['G', '', 297.3, 57.05, None, ''])
        p = tmp_path / 'mat.xlsx'
        wb.save(str(p))
        rows = load_materials_file(p)
        assert len(rows) == 1
        assert rows[0]['token'] == 'G'

    def test_unknown_extension_tries_csv(self, tmp_path):
        f = tmp_path / 'mat.txt'
        f.write_text(
            'ResidueCode,ProtectionGroup,FmocMW_g_mol,FreeAA_MW_g_mol,Density_g_mL,Notes\n'
            'A,,311.3,71.08,,\n'
        )
        rows = load_materials_file(f)
        assert len(rows) == 1
