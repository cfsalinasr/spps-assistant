"""Tests for infrastructure/xlsx_generator.py."""

from pathlib import Path

import openpyxl
import pytest

from spps_assistant.domain.models import MaterialsRow
from spps_assistant.infrastructure.xlsx_generator import generate_materials_xlsx


def _make_row(**kwargs) -> MaterialsRow:
    """Build a MaterialsRow with sensible test defaults."""
    defaults = dict(
        token='A', protection='', fmoc_mw=311.3,
        mmol_needed=0.09, mass_mg=28.02, stock_conc=0.5,
        volume_ml=0.600, notes='Fmoc-Ala-OH', formula='V = ...',
    )
    defaults.update(kwargs)
    return MaterialsRow(**defaults)


# ── file creation ─────────────────────────────────────────────────────────────

class TestFileCreation:
    def test_creates_file(self, tmp_path):
        """XLSX file is created at the given path."""
        path = tmp_path / 'out.xlsx'
        generate_materials_xlsx(path, 'TestSynth', [_make_row()])
        assert path.exists()

    def test_creates_parent_dirs(self, tmp_path):
        """Generator creates intermediate directories as needed."""
        path = tmp_path / 'sub' / 'dir' / 'out.xlsx'
        generate_materials_xlsx(path, 'TestSynth', [_make_row()])
        assert path.exists()

    def test_empty_rows_creates_file(self, tmp_path):
        """Generator creates a valid XLSX even with no data rows."""
        path = tmp_path / 'out.xlsx'
        generate_materials_xlsx(path, 'TestSynth', [])
        assert path.exists()

    def test_file_is_valid_xlsx(self, tmp_path):
        """Output file is a valid openpyxl workbook."""
        path = tmp_path / 'out.xlsx'
        generate_materials_xlsx(path, 'TestSynth', [_make_row()])
        wb = openpyxl.load_workbook(str(path))
        assert wb is not None


# ── sheet structure ───────────────────────────────────────────────────────────

class TestSheetStructure:
    def _load_ws(self, tmp_path, rows=None, name='TestSynth'):
        """Generate XLSX and return its active worksheet."""
        path = tmp_path / 'out.xlsx'
        if rows is None:
            rows = [_make_row()]
        generate_materials_xlsx(path, name, rows)
        wb = openpyxl.load_workbook(str(path))
        return wb.active

    def test_sheet_title_truncated_to_31_chars(self, tmp_path):
        """Sheet name is capped at 31 characters."""
        long_name = 'A' * 50
        ws = self._load_ws(tmp_path, name=long_name)
        assert len(ws.title) <= 31

    def test_title_row_has_synthesis_name(self, tmp_path):
        """Title row A1 contains the synthesis name."""
        ws = self._load_ws(tmp_path, name='MySynth')
        assert 'MySynth' in str(ws['A1'].value)

    def test_header_row_has_nine_columns(self, tmp_path):
        """Header row (row 2) has nine non-empty column headers."""
        ws = self._load_ws(tmp_path)
        headers = [ws.cell(row=2, column=c).value for c in range(1, 10)]
        assert all(h is not None for h in headers)

    def test_header_residue_column(self, tmp_path):
        """First column header is 'Residue'."""
        ws = self._load_ws(tmp_path)
        assert ws.cell(row=2, column=1).value == 'Residue'

    def test_data_starts_at_row_3(self, tmp_path):
        """Data rows begin at row 3."""
        rows = [_make_row(token='A'), _make_row(token='G', fmoc_mw=297.3)]
        ws = self._load_ws(tmp_path, rows=rows)
        assert ws.cell(row=3, column=1).value == 'A'
        assert ws.cell(row=4, column=1).value == 'G'

    def test_freeze_panes_set(self, tmp_path):
        """Freeze panes are set so headers remain visible when scrolling."""
        ws = self._load_ws(tmp_path)
        assert ws.freeze_panes is not None


# ── data values ───────────────────────────────────────────────────────────────

class TestDataValues:
    def _load_data_row(self, tmp_path, row: MaterialsRow):
        """Generate XLSX with one row and return the cell values for that row."""
        path = tmp_path / 'out.xlsx'
        generate_materials_xlsx(path, 'T', [row])
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        return [ws.cell(row=3, column=c).value for c in range(1, 10)]

    def test_token_written(self, tmp_path):
        """Token value is written to column 1."""
        vals = self._load_data_row(tmp_path, _make_row(token='C(Trt)'))
        assert vals[0] == 'C(Trt)'

    def test_fmoc_mw_written(self, tmp_path):
        """Fmoc MW is written to column 3."""
        vals = self._load_data_row(tmp_path, _make_row(fmoc_mw=585.7))
        assert vals[2] == pytest.approx(585.7)

    def test_mass_mg_written(self, tmp_path):
        """Mass in mg is written to column 5."""
        vals = self._load_data_row(tmp_path, _make_row(mass_mg=52.7))
        assert vals[4] == pytest.approx(52.7)

    def test_volume_ml_written(self, tmp_path):
        """Volume in mL is written to column 7."""
        vals = self._load_data_row(tmp_path, _make_row(volume_ml=1.234))
        assert vals[6] == pytest.approx(1.234)

    def test_notes_written(self, tmp_path):
        """Notes text is written to column 9."""
        vals = self._load_data_row(tmp_path, _make_row(notes='My note'))
        assert vals[8] == 'My note'

    def test_multiple_rows(self, tmp_path):
        """Multiple rows are all written in order."""
        rows = [
            _make_row(token='A', fmoc_mw=311.3),
            _make_row(token='G', fmoc_mw=297.3),
            _make_row(token='W', fmoc_mw=496.6),
        ]
        path = tmp_path / 'out.xlsx'
        generate_materials_xlsx(path, 'T', rows)
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        tokens = [ws.cell(row=r, column=1).value for r in range(3, 6)]
        assert tokens == ['A', 'G', 'W']
